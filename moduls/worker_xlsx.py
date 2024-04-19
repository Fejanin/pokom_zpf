import openpyxl


class XLSX_Data:
    '''Хранит данные по одной строке:
- код единицы продаж;
- код продукта;
- номер варианта;
- штрих-код;
- наименование;
- количество в заказе;
- и порядковый номер строки, из которой были записаны данные.'''
    START_SLICE = 0
    END_SLICE = 4
    NAME_PRODUCT = 14 # O - колонка
    NUM_ORDER = 22 # W
    CONTROL_KG = 21 # V
    NAME_COL = 'W'
    def __new__(cls, data, number, *args, **kwargs):
        res = list(map(lambda x: x.value, data))
        if cls.control_line(res):
            return super().__new__(cls)


    def __init__(self, data, number):
        self.sales_unit_code, self.product_code, self.option_number, self.barcode, self.product, self.num_order = self.read_line(data)
        self.num_row = number


    def read_line(self, data):
        '''Считываем целую строку из таблицы и выбираем те данные, которые нам нужны.'''
        res = list(map(lambda x: x.value, data))
        if self.control_line(res):
            return res[self.START_SLICE: self.END_SLICE] + [res[self.NAME_PRODUCT], res[self.NUM_ORDER]]


    def __str__(self):
        return f'Строка - {self.num_row}; наименование - {self.product}'


    def __eq__(self, obj):
        return self.sales_unit_code == obj.sales_unit_code and self.product_code == obj.product_code and self.option_number == obj.option_number and self.barcode == obj.barcode


    def __setattr__(self, key, value):
        '''Удаляем пробелы из начала и окончания строки.'''
        if type(value) is str:
            value = value.strip()
        object.__setattr__(self, key, value)


    @classmethod
    def control_line(cls, data):
        return all(data[cls.START_SLICE: cls.END_SLICE]) and data[cls.CONTROL_KG] == 'кор'


class POKOM_Reader:
    '''Скачивает данные из file.
Формирует список объектов XLSX_Data. Предназначен для работы со старыми версиями заказников.'''
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.active
        self.all_rows = []
        self.read()


    def read(self):
        '''Создаем объект класса XLSX_Data и добавляем его в список all_rows'''
        a = self.ws.iter_rows(1, self.ws.max_row + 1)
        for num, i in enumerate(a, 1):
            obj = XLSX_Data(i, num)
            if obj:
                self.all_rows.append(obj)


    def __call__(self):
        return round(sum([i.num_order for i in self.all_rows]), 3)


class OneC:
    '''Скачивает данные из file.
Формирует список объектов XLSX_Data. Предназначен для работы с любыми xslx-файлами. Главные условия:
- в первом столбце находятся наименования СКЮ;
- во втором столбце их заказ в кг!'''
    def __init__(self, file):
        self.error = []
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.active
        self.keys_values = self.read()
        self.translater_data = self.find_keys()
        self.all_rows = self.translater()  # хранит список всех объектов OneCData
        

    def read(self):
        '''Считывает данные из файла, хранящего заказ. Наименования должны соответствовать СКЮ 1С!!!
Получает следующие данные:
    - Название СКЮ (в дальнейшем будет ключом);
        - количесто шт/кг из заказа
        - номер строки.'''
        a = self.ws.iter_rows(1, self.ws.max_row)
        res = {}
        for num, i in enumerate(a, 1):
            if i:
                data = list(map(lambda x: x.value if type(x.value) in (int, float) else x.value.strip() if type(x.value) is str else None, i))
                if data[1] and data[1] > 0:
                    key = data[0].strip()
                    if data[0].strip() in res:
                        self.error.append(f'Объект - {data[0].strip()} в количестве {data[1]} не добавлен, т.к. в файле заказа данная СКЮ уже было в количестве {res[key]["value"]} в строке {res[key]["row"]}.')
                        continue
                    res[data[0].strip()] = {'value': data[1], 'row': num}
        return res


    def find_keys(self):
        '''Открываем файл "переводчик" и подтягиваем из него все известные ключи и их значения.'''
        file = r'moduls\1С.xlsx' # файл с ключами для поиска
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        a = ws.iter_rows(1, ws.max_row + 1)
        res = {}
        for i in a:
            time_data = list(map(lambda x: x.value, i))
            if time_data[0] and time_data[0] != '1С':
                res[time_data[0].strip()] = time_data[1:5]
        return res


    def translater(self):
        '''На основе скаченных ключей, создаем объекты класса OneCData и добавляем их в список.'''
        res = []
        for i in self.keys_values:
            if i in self.translater_data:
                res.append(OneCData(self.translater_data[i] + [i, self.keys_values[i]['value']], self.keys_values[i]['row']))
            else:
                res.append(OneCData([None for _ in range(4)]+ [i, self.keys_values[i]['value']], self.keys_values[i]['row']))
        return res


    def __call__(self):
        return round(sum([i.num_order for i in self.all_rows]), 3)


class OneCData:
    '''Создаем класс со структурой подобной классу XLSX_Data, для удобного их сравнения.'''
    def __init__(self, data, number):
        self.sales_unit_code, self.product_code, self.option_number, self.barcode, self.product, self.num_order = data
        self.num_row = number


    def __str__(self):
        return f'Строка - {self.num_row}; наименование - {self.product}'


    def __eq__(self, obj):
        return self.sales_unit_code == obj.sales_unit_code and self.product_code == obj.product_code and self.option_number == obj.option_number and self.barcode == obj.barcode


    def __setattr__(self, key, value):
        '''Удаляем пробелы из начала и окончания строки.'''
        if type(value) is str:
            value = value.strip()
        object.__setattr__(self, key, value)


class POKOM_Rewriter:
    '''Основной класс. Создает списки с данными из читаемого и записываемого файла, проводит их сравнения и записывает совпавшие данные.
Создает файл REPORT.txt с отчетом о проделанной работе.'''
    def __init__(self, read_name_file, write_name_file, flag_pocom = False):
        self.tracker = Tracker()
        if not flag_pocom: # файл ПОКОМ
            self.read_file = POKOM_Reader(read_name_file)
        else:
            if flag_pocom == '1С': # "С" - символ кирилицей!!!!
                self.read_file = OneC(read_name_file)
                self.tracker.error += self.read_file.error
        self.write_file = POKOM_Reader(write_name_file)
        for i in self.read_file.all_rows:
            flag = False
            obj1 = str(i)
            for j in self.write_file.all_rows:
                obj2 = str(j)
                if i == j and i.num_order:
                    flag = True
                    if self.write_file.ws[j.NAME_COL + str(j.num_row)].value:
                        self.tracker.error.append(f'Объект - {obj1} в количестве {i.num_order} не добавлен, т.к. ячейка {j.NAME_COL + str(j.num_row)} уже содержит значение {self.write_file.ws[j.NAME_COL + str(j.num_row)].value}.')
                        continue
                    self.write(j, i.num_order)
                    self.tracker.message.append(f'Перенос данных из {obj1} ==> \n\t{obj2}, в количестве {i.num_order}')
                    self.tracker.wight2 += i.num_order
                    continue
            if not flag and i.num_order: # ТРЕБУЕТСЯ ПРОВЕРИТЬ ПРАВИЛЬНОСТЬ ПРОВЕРКИ!!!
                self.tracker.error.append(f'Объект - {obj1} в количестве {i.num_order} НЕ НАЙДЕН.')
        self.read_file.wb.save(read_name_file)
        self.write_file.wb.save(write_name_file)
        self.tracker.wight1 = self.read_file()
        self.tracker.create_file()


    def write(self, obj, num):
        '''Сохранение данных в бланк заказа.'''
        self.write_file.ws[obj.NAME_COL + str(obj.num_row)] = num


class Tracker:
    '''Создает текстовый файл с результатами переноса данных (в т.ч. и ошибками)'''
    def __init__(self):
        self.error = []
        self.message = []
        self.wight1 = 0
        self.wight2 = 0


    def create_file(self):
        '''Создание файла с отчетом о проделанной работе.'''
        with open('REPORT.txt', 'w') as f:
            if self.error:
                f.write(f'Обнаружены следующие ошибки:\n')
                for i in self.error:
                    f.write(i + '\n')
                f.write('#' * 50 + '\n\n')
            for i in self.message:
                f.write(i + '\n')
            f.write(f'\nКоличество коробок с  продукцией из первого файла составляет: {self.wight1} шт.\n')
            f.write(f'Количество коробок продукции из второго фала составляет: {round(self.wight2, 3)} шт.\n')

