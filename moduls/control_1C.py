import openpyxl


file = '1С.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb.active
control_dict = {}
a = ws.iter_rows(1, ws.max_row + 1)
for num, i in enumerate(a, 1):
    val = i[0].value
    if val:
        val = val.strip()
        d = control_dict.setdefault(val, []).append(num)
res = []
for i in control_dict:
    if len(control_dict[i]) > 1:
        res.append(f'{i} - {control_dict[i]}')
if res:
    print(*[i for i in res], sep='/n')
else:
    print('Задвоенных позиций нет.')
input('Нажмите ENTER')
